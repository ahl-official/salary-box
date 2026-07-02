from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, date
import os, calendar
from models.sheets import ReportSheet, EmployeeSheet, AttendanceSheet, SettingsSheet, BranchSheet, HolidaySheet
from utils.datetime_utils import now_local, parse_timestamp
from utils.shift_utils import parse_working_days, overtime_hours
from utils.branch_settings import effective_settings
from utils.late_policy import build_monthly_late_summary
from utils.holiday_utils import holiday_on_date
from utils.auth import get_current_user, require_admin

router = APIRouter()
REPORTS_DIR = "generated_reports"
os.makedirs(REPORTS_DIR, exist_ok=True)

class ReportRequest(BaseModel):
    type: str
    month: Optional[str] = None   # YYYY-MM
    branch: Optional[str] = None
    department: Optional[str] = None

@router.get("/")
def list_reports(current_user: dict = Depends(require_admin)):
    return ReportSheet.all()

@router.post("/generate")
def generate_report(req: ReportRequest, current_user: dict = Depends(require_admin)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    s = SettingsSheet.get_all()

    if req.month:
        year, month = map(int, req.month.split("-"))
    else:
        local_today = now_local(s)
        year, month = local_today.year, local_today.month

    days_in_month = calendar.monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date   = f"{year}-{month:02d}-{days_in_month:02d}"

    employees = EmployeeSheet.all(
        department=req.department if req.department and req.department != "All" else "",
        branch=req.branch if req.branch and req.branch != "All" else "",
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    accent    = "00C896"
    dark_bg   = "1A1A1A"
    header_fill   = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
    subhdr_fill   = PatternFill(start_color=dark_bg, end_color=dark_bg, fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True, name="Calibri")
    black_bold = Font(color="000000", bold=True, name="Calibri")
    thin = Side(border_style="thin", color="333333")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left", vertical="center")

    days  = list(range(1, days_in_month + 1))
    dnames = [calendar.day_abbr[calendar.weekday(year, month, d)] for d in days]

    fixed_cols = ["#","Name","Designation","Department","Branch","Present","Absent","Week Off","Late Days","Half-Day Ded.","Total Hrs","OT Hrs"]
    all_cols   = fixed_cols + [f"{d}\n{dnames[i]}" for i, d in enumerate(days)]

    # ---- Header rows ----
    r = 1
    ws.merge_cells(f"A{r}:Z{r}")
    ws[f"A{r}"] = s.get("company_name", "Company")
    ws[f"A{r}"].font = Font(bold=True, size=16, name="Calibri", color="FFFFFF")
    ws[f"A{r}"].alignment = ctr
    ws[f"A{r}"].fill = subhdr_fill

    r += 1
    ws.merge_cells(f"A{r}:Z{r}")
    ws[f"A{r}"] = f"{req.type.replace('_',' ').title()}  |  {req.month or date.today().strftime('%Y-%m')}  |  {req.branch or 'All'}  |  {req.department or 'All'}"
    ws[f"A{r}"].font = Font(bold=True, size=11, color=accent, name="Calibri")
    ws[f"A{r}"].alignment = ctr
    ws[f"A{r}"].fill = subhdr_fill

    r += 1
    ws.merge_cells(f"A{r}:Z{r}")
    ws[f"A{r}"] = f"Address: {s.get('address','')}   GSTIN: {s.get('gstin','')}"
    ws[f"A{r}"].font = Font(size=9, name="Calibri", color="888888")
    ws[f"A{r}"].alignment = ctr

    r += 2
    for ci, col in enumerate(all_cols, 1):
        cell = ws.cell(row=r, column=ci, value=col)
        cell.font  = black_bold
        cell.fill  = header_fill
        cell.alignment = ctr
        cell.border = border
    ws.row_dimensions[r].height = 30

    r += 1
    for idx, emp in enumerate(employees, 1):
        punches = AttendanceSheet.for_employee_month(emp["id"], year, month)
        day_map: dict = {}
        for p in punches:
            try:
                d = parse_timestamp(str(p["timestamp"]), s).date().isoformat()
            except ValueError:
                continue
            if d not in day_map:
                day_map[d] = {"in": None, "out": None}
            if p["punch_type"] == "in"  and not day_map[d]["in"]:
                day_map[d]["in"]  = p["timestamp"]
            if p["punch_type"] == "out":
                day_map[d]["out"] = p["timestamp"]

        present = 0; total_hours = 0.0; total_ot = 0.0
        statuses = {}
        emp_branch = BranchSheet.get_by_name(emp.get("branch", "")) if emp.get("branch") else None
        emp_settings = effective_settings(s, emp_branch)
        working_days = parse_working_days(emp_settings)
        late_summary = build_monthly_late_summary(punches, emp_settings, year, month)
        month_holidays = HolidaySheet.all(month=f"{year}-{month:02d}")
        for i, d in enumerate(days):
            key = f"{year}-{month:02d}-{d:02d}"
            dn  = dnames[i]
            day_off = dn not in working_days
            info = day_map.get(key)
            emp_holiday = holiday_on_date(month_holidays, emp["id"], key)
            if info and info["in"]:
                present += 1
                if info["out"]:
                    day_hours = (parse_timestamp(info["out"], emp_settings) - parse_timestamp(info["in"], emp_settings)).total_seconds()/3600
                    total_hours += day_hours
                    day_ot = overtime_hours(day_hours, emp_settings)
                    total_ot += day_ot
                day_late = late_summary["days"].get(key)
                if day_late and day_late.get("is_late"):
                    if day_late.get("half_day_deduction"):
                        statuses[key] = f"L-{day_late['late_minutes']}"
                    elif day_late.get("forgiven"):
                        statuses[key] = f"L*{day_late['late_minutes']}"
                    else:
                        statuses[key] = f"L{day_late['late_minutes']}"
                elif info.get("out"):
                    day_ot = overtime_hours(
                        (parse_timestamp(info["out"], emp_settings) - parse_timestamp(info["in"], emp_settings)).total_seconds()/3600,
                        emp_settings,
                    )
                    statuses[key] = f"P+{day_ot:.1f}" if day_ot > 0 else "P"
                else:
                    statuses[key] = "P"
            elif emp_holiday:
                statuses[key] = "H"
            elif day_off:
                statuses[key] = "WO"
            else:
                statuses[key] = "A"

        absent   = sum(1 for v in statuses.values() if v == "A")
        week_off = sum(1 for v in statuses.values() if v == "WO")

        row_data = [idx, emp["name"], emp.get("designation",""), emp.get("department",""), emp.get("branch",""),
                    present, absent, week_off,
                    late_summary["late_days_count"], late_summary["half_day_deductions"],
                    f"{total_hours:.1f}h", f"{total_ot:.1f}h"]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = border
            cell.alignment = lft if ci == 2 else ctr

        for i, d in enumerate(days):
            key = f"{year}-{month:02d}-{d:02d}"
            st  = statuses.get(key, "-")
            ci  = len(fixed_cols) + i + 1
            cell = ws.cell(row=r, column=ci, value=st)
            cell.border = border
            cell.alignment = ctr
            if st == "P" or str(st).startswith("P+"):
                cell.font = Font(color="00C896", bold=True, name="Calibri")
            elif st == "A":
                cell.font = Font(color="FF4444", bold=True, name="Calibri")
            elif st == "WO":
                cell.font = Font(color="888888", name="Calibri")
            elif st == "H":
                cell.font = Font(color="6FA8FF", bold=True, name="Calibri")
        r += 1

    # Legend
    r += 1
    ws.merge_cells(f"A{r}:Z{r}")
    ws[f"A{r}"] = "P = Present  |  P+X = OT  |  H = Holiday  |  L* / L- = Late  |  A = Absent  |  WO = Week Off"
    ws[f"A{r}"].font = Font(italic=True, size=9, color="666666", name="Calibri")

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    for ci in range(4, len(fixed_cols)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    for ci in range(len(fixed_cols)+1, len(all_cols)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 6
    ws.freeze_panes = "F5"

    ts       = now_local(s).strftime("%Y%m%d_%H%M%S")
    rname    = f"{req.type}_{req.month or now_local(s).strftime('%Y-%m')}_{ts}"
    filename = f"{rname}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb.save(filepath)

    report = ReportSheet.create(
        name=rname, rtype=req.type,
        month=req.month or now_local(s).strftime("%Y-%m"),
        branch=req.branch or "All",
        department=req.department or "All",
        file_path=filepath,
    )
    return {"success": True, "report": report, "filename": filename}

@router.get("/download/{report_id}")
def download_report(report_id: int, current_user: dict = Depends(require_admin)):
    report = ReportSheet.get_by_id(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    if not os.path.exists(report["file_path"]):
        raise HTTPException(status_code=404, detail="File not found on disk")
    return FileResponse(
        report["file_path"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(report["file_path"]),
    )

@router.delete("/{report_id}")
def delete_report(report_id: int, current_user: dict = Depends(require_admin)):
    report = ReportSheet.get_by_id(report_id)
    ReportSheet.delete(report_id)
    if report and os.path.exists(report.get("file_path", "")):
        os.remove(report["file_path"])
    return {"success": True}
